"""
LexAid Selenium E2E — Multi-Sheet Excel Report Generator
=========================================================
Generates:
  - Automation_Test_Report.xlsx  (6 sheets)
  - Passed_Test_Cases.xlsx
  - Failed_Test_Cases.xlsx
  - Summary_Report.xlsx
"""

import os
import openpyxl
from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side)
from openpyxl.utils import get_column_letter
from datetime import datetime

# ── Color palette ──────────────────────────────────────
NAVY        = "1F4E78"
WHITE       = "FFFFFF"
PASS_GREEN  = "C6E0B4"
FAIL_RED    = "FFC7CE"
SKIP_YELLOW = "FFEB9C"
HEADER_BLUE = "2F5496"
LIGHT_GREY  = "F2F2F2"

def _hdr_font():      return Font(bold=True, color=WHITE, size=11)
def _hdr_fill(hex):   return PatternFill("solid", fgColor=hex)
def _border():
    s = Side(border_style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)
def _center():        return Alignment(horizontal="center", vertical="center", wrap_text=True)


def _apply_header(ws, columns: list, fill_color: str = HEADER_BLUE):
    ws.append([c["header"] for c in columns])
    for i, col in enumerate(columns, 1):
        ws.column_dimensions[get_column_letter(i)].width = col.get("width", 20)
    row = ws[1]
    for cell in row:
        cell.font   = _hdr_font()
        cell.fill   = _hdr_fill(fill_color)
        cell.alignment = _center()
        cell.border = _border()
    ws.row_dimensions[1].height = 28


def _color_row(ws, row_num: int, fill_color: str):
    fill = PatternFill("solid", fgColor=fill_color)
    for cell in ws[row_num]:
        cell.fill   = fill
        cell.border = _border()
        cell.alignment = Alignment(vertical="center", wrap_text=True)


MAIN_COLUMNS = [
    {"header": "Test ID",            "width": 18},
    {"header": "Module",             "width": 25},
    {"header": "Test Name",          "width": 45},
    {"header": "Priority",           "width": 12},
    {"header": "Status",             "width": 14},
    {"header": "Duration (ms)",      "width": 16},
    {"header": "Failure Reason",     "width": 45},
    {"header": "Screenshot",         "width": 35},
]


def generate_excel_reports(results: list, target_dir: str, build_number: str = ""):
    os.makedirs(target_dir, exist_ok=True)

    passed  = [r for r in results if r["status"] == "PASSED"]
    failed  = [r for r in results if r["status"] == "FAILED"]
    skipped = [r for r in results if r["status"] == "SKIPPED"]
    total   = len(results)
    pass_rate = (len(passed) / total * 100) if total else 0

    # ── Module pass rate map ──────────────────────────
    module_map: dict = {}
    for r in results:
        m = r["module"]
        if m not in module_map:
            module_map[m] = {"total": 0, "passed": 0, "failed": 0}
        module_map[m]["total"]  += 1
        if r["status"] == "PASSED":
            module_map[m]["passed"] += 1
        elif r["status"] == "FAILED":
            module_map[m]["failed"] += 1

    # ═══════════════════════════════════════════════════
    # 1. Automation_Test_Report.xlsx  (6 sheets)
    # ═══════════════════════════════════════════════════
    wb = openpyxl.Workbook()

    # Sheet 1 — Executed Test Cases
    ws1 = wb.active
    ws1.title = "Executed Test Cases"
    _apply_header(ws1, MAIN_COLUMNS, NAVY)
    for r in results:
        ws1.append([
            r["id"], r["module"], r["name"], r.get("priority","P2"),
            r["status"], r.get("duration_ms", 0),
            r.get("reason") or "—",
            r.get("screenshot") or "—"
        ])
        row_n = ws1.max_row
        fill  = PASS_GREEN if r["status"]=="PASSED" else FAIL_RED if r["status"]=="FAILED" else SKIP_YELLOW
        _color_row(ws1, row_n, fill)

    # Sheet 2 — Passed Tests
    ws2 = wb.create_sheet("Passed Tests")
    _apply_header(ws2, MAIN_COLUMNS, "375623")
    for r in passed:
        ws2.append([r["id"], r["module"], r["name"], r.get("priority","P2"),
                    r["status"], r.get("duration_ms",0), "—", "—"])
        _color_row(ws2, ws2.max_row, PASS_GREEN)

    # Sheet 3 — Failed Tests
    ws3 = wb.create_sheet("Failed Tests")
    _apply_header(ws3, MAIN_COLUMNS, "9C0006")
    for r in failed:
        ws3.append([r["id"], r["module"], r["name"], r.get("priority","P2"),
                    r["status"], r.get("duration_ms",0),
                    r.get("reason") or "Assertion failed",
                    r.get("screenshot") or "—"])
        _color_row(ws3, ws3.max_row, FAIL_RED)

    # Sheet 4 — Skipped Tests
    ws4 = wb.create_sheet("Skipped Tests")
    _apply_header(ws4, [
        {"header": "Test ID", "width": 18}, {"header": "Module","width":25},
        {"header": "Test Name","width":45}, {"header": "Skip Reason","width":35}
    ], "7F6000")
    for r in skipped:
        ws4.append([r["id"], r["module"], r["name"], r.get("reason","Feature not enabled")])
        _color_row(ws4, ws4.max_row, SKIP_YELLOW)

    # Sheet 5 — Execution Metrics
    ws5 = wb.create_sheet("Execution Metrics")
    _apply_header(ws5, [{"header":"Metric","width":35},{"header":"Value","width":25}], HEADER_BLUE)
    metrics = [
        ("Build Number", build_number),
        ("Execution Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Target URL", os.environ.get("BASE_URL","https://suvan789.github.io/lexaid/")),
        ("Total Test Cases", total),
        ("Passed", len(passed)),
        ("Failed", len(failed)),
        ("Skipped", len(skipped)),
        ("Pass Rate (%)", f"{pass_rate:.2f}%"),
        ("Pass Threshold (%)", "95.00%"),
        ("Threshold Status", "✅ MET" if pass_rate >= 95 else "❌ NOT MET"),
        ("Browser", "Headless Chrome"),
        ("Selenium Version", "4.x"),
    ]
    for m, v in metrics:
        ws5.append([m, str(v)])
        _color_row(ws5, ws5.max_row, LIGHT_GREY)

    # Sheet 6 — Defect Summary
    ws6 = wb.create_sheet("Defect Summary")
    _apply_header(ws6, [
        {"header":"Defect ID","width":15},{"header":"Test ID","width":18},
        {"header":"Module","width":25},{"header":"Priority","width":12},
        {"header":"Severity","width":15},{"header":"Summary","width":50}
    ], "843C0C")
    for idx, r in enumerate(failed, 101):
        ws6.append([
            f"DEF-{idx}", r["id"], r["module"], r.get("priority","P2"),
            "HIGH" if r.get("priority")=="P1" else "MEDIUM",
            f"{r['module']}: {r.get('reason','Assertion failed')[:80]}"
        ])
        _color_row(ws6, ws6.max_row, FAIL_RED)

    wb.save(os.path.join(target_dir, "Automation_Test_Report.xlsx"))

    # ═══════════════════════════════════════════════════
    # 2. Passed_Test_Cases.xlsx
    # ═══════════════════════════════════════════════════
    wb2 = openpyxl.Workbook()
    ws = wb2.active; ws.title = "Passed"
    _apply_header(ws, MAIN_COLUMNS, "375623")
    for r in passed:
        ws.append([r["id"],r["module"],r["name"],r.get("priority","P2"),"PASSED",r.get("duration_ms",0),"—","—"])
        _color_row(ws, ws.max_row, PASS_GREEN)
    wb2.save(os.path.join(target_dir, "Passed_Test_Cases.xlsx"))

    # ═══════════════════════════════════════════════════
    # 3. Failed_Test_Cases.xlsx
    # ═══════════════════════════════════════════════════
    wb3 = openpyxl.Workbook()
    ws = wb3.active; ws.title = "Failed"
    _apply_header(ws, MAIN_COLUMNS, "9C0006")
    for r in failed:
        ws.append([r["id"],r["module"],r["name"],r.get("priority","P2"),"FAILED",
                   r.get("duration_ms",0),r.get("reason","Assertion failed"),r.get("screenshot","—")])
        _color_row(ws, ws.max_row, FAIL_RED)
    wb3.save(os.path.join(target_dir, "Failed_Test_Cases.xlsx"))

    # ═══════════════════════════════════════════════════
    # 4. Summary_Report.xlsx
    # ═══════════════════════════════════════════════════
    wb4 = openpyxl.Workbook()
    ws = wb4.active; ws.title = "Summary"
    _apply_header(ws, [{"header":"Metric","width":35},{"header":"Value","width":25}], NAVY)
    for m, v in metrics:
        ws.append([m, str(v)])
    wb4.save(os.path.join(target_dir, "Summary_Report.xlsx"))

    print(f"✅ Excel reports generated in: {target_dir}")
    return {"passed": len(passed), "failed": len(failed), "skipped": len(skipped),
            "total": total, "pass_rate": pass_rate}

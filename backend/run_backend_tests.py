import pytest
import openpyxl
from openpyxl.styles import PatternFill, Font
from datetime import datetime
import os

class ExcelReporter:
    def __init__(self):
        self.passedTests = []
        self.failedTests = []
        self.executionLog = []

    def pytest_runtest_logreport(self, report):
        if report.when == "call":
            duration = round(report.duration, 2)
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            # Extract category and test name
            node_parts = report.nodeid.split("::")[-1].split("_")
            category = "Unknown"
            if len(node_parts) > 4:
                category = node_parts[3].capitalize()
            test_name = report.nodeid.split("::")[-1]

            if report.passed:
                self.passedTests.append({
                    "category": category,
                    "testName": test_name,
                    "time": duration
                })
                self.executionLog.append({
                    "timestamp": timestamp,
                    "level": "INFO",
                    "message": f"[{category}] {test_name} -> PASSED in {duration}s"
                })
            elif report.failed:
                error_msg = str(report.longrepr)
                self.failedTests.append({
                    "category": category,
                    "testName": test_name,
                    "error": error_msg
                })
                self.executionLog.append({
                    "timestamp": timestamp,
                    "level": "ERROR",
                    "message": f"[{category}] {test_name} -> FAILED: {error_msg.splitlines()[-1]}"
                })

def main():
    reporter = ExcelReporter()
    print("Starting execution of 110 REAL FastAPI backend tests...")
    
    # Run pytest programmatically
    failures = pytest.main(["tests/test_api.py", "-v", "--tb=short", "-W", "ignore"], plugins=[reporter])
    
    print("\nGenerating Excel Report...")
    wb = openpyxl.Workbook()
    
    # Header Styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    pass_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    fail_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
    
    # 1. Summary
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Metric", "Value"])
    for cell in ws_summary[1]:
        cell.fill = header_fill
        cell.font = header_font
        
    total_passed = len(reporter.passedTests)
    total_failed = len(reporter.failedTests)
    ws_summary.append(["Total Tests", total_passed + total_failed])
    ws_summary.append(["Passed", total_passed])
    ws_summary.append(["Failed", total_failed])
    ws_summary.append(["Status", "FAILED" if total_failed > 0 else "PASSED"])
    
    # 2. Passed Tests
    ws_passed = wb.create_sheet("Passed Tests")
    ws_passed.append(["No.", "Category", "Test Name", "Time (sec)", "Status"])
    for cell in ws_passed[1]:
        cell.fill = header_fill
        cell.font = header_font
        
    for idx, test in enumerate(reporter.passedTests, 1):
        ws_passed.append([idx, test["category"], test["testName"], test["time"], "PASSED"])
        for cell in ws_passed[ws_passed.max_row]:
            cell.fill = pass_fill

    # 3. Failed Tests
    ws_failed = wb.create_sheet("Failed Tests")
    ws_failed.append(["No.", "Category", "Test Name", "Error Details", "Status"])
    for cell in ws_failed[1]:
        cell.fill = header_fill
        cell.font = header_font
        
    for idx, test in enumerate(reporter.failedTests, 1):
        ws_failed.append([idx, test["category"], test["testName"], test["error"], "FAILED"])
        for cell in ws_failed[ws_failed.max_row]:
            cell.fill = fail_fill
            
    # 4. Execution Log
    ws_log = wb.create_sheet("Execution Log")
    ws_log.append(["Timestamp", "Level", "Message"])
    for cell in ws_log[1]:
        cell.fill = header_fill
        cell.font = header_font
        
    for log in reporter.executionLog:
        ws_log.append([log["timestamp"], log["level"], log["message"]])
        fill = pass_fill if log["level"] == "INFO" else fail_fill
        for cell in ws_log[ws_log.max_row]:
            cell.fill = fill
            
    # 5. Test Details
    wb.create_sheet("Test Details")

    filename = f"Backend_Test_Report_LexAid_{datetime.now().strftime('%Y-%m-%dT%H-%M-%S')}.xlsx"
    wb.save(filename)
    
    print("======================================================")
    print(f"All Tests Finished! Executed {total_passed + total_failed} real API validations.")
    print(f"Generated true Excel report: {filename}")
    print("======================================================")

if __name__ == "__main__":
    main()
